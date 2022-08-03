# coding=utf-8
import logging
import os
import config
import catalog.models as m

from openpyxl import load_workbook
from openpyxl.utils import quote_sheetname, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font

from catalog import services, utils
from flask import current_app as app, request

__author__ = 'Linh.VH'
_logger = logging.getLogger(__name__)


def _add_data_validation_list(ws, data_length, char_from, char_to):
    """
    Add data validation dropdown list for SanPham sheet (with data in DuLieuMau sheet)
    :param:
    :return:
    """
    data_val = DataValidation(type="list", showErrorMessage=False,
                              formula1='{sheet_name}!${letter}$2:${letter}${len}'.format(
                                  sheet_name=quote_sheetname('DuLieuMau'), letter=char_from, len=data_length + 1))
    ws.add_data_validation(data_val)
    if type(char_to) is list:
        for ct in char_to:
            data_val.add('{letter}9:{letter}1048576'.format(letter=ct))
    else:
        data_val.add('{letter}9:{letter}1048576'.format(letter=char_to))


def get_data_for_excel_sheet(attribute_set_id):
    """

    :param attribute_set_id:
    :return:
    """
    attributes_data = {}
    if attribute_set_id is not None:
        attributes_data = services.attribute_set.get_attribute_set_content(
            attribute_set_id
        )
    extra_data = services.extra.get_extra_info(req_restrict_data='1')
    extra_data['types'] = ['Bundle', 'Simple']
    extra_data['categories'] = utils.filter_active_leaf_categories(
        extra_data['categories']
    )
    extra_data['categories'].sort(key=lambda x: utils.normalized(x))
    extra_data['sale_categories'] = utils.filter_active_leaf_categories(
        extra_data['sale_categories']
    )
    extra_data['sale_categories'].sort(key=lambda x: utils.normalized(x))
    extra_data['brands'] = [brand.name for brand in extra_data['brands']]
    extra_data['brands'].sort(key=lambda x: utils.normalized(x))
    extra_data['colors'] = [color.name for color in extra_data['colors']]
    extra_data['colors'].sort(key=lambda x: utils.normalized(x))
    extra_data['units'] = [unit.name for unit in extra_data['units']]
    extra_data['units'].sort(key=lambda x: utils.normalized(x))
    extra_data['product_types'] = [product_type.name for product_type
                                   in extra_data['product_types']]
    extra_data['product_types'].sort(key=lambda x: utils.normalized(x))
    extra_data['objectives'] = [objective.name for objective
                                in extra_data['objectives']]
    extra_data['objectives'].sort(key=lambda x: utils.normalized(x))
    extra_data['channels'] = [channel.name for channel
                              in extra_data['channels']]
    extra_data['channels'].sort(key=lambda x: utils.normalized(x))
    extra_data['editing_status'] = [editing_status.name for editing_status
                                    in extra_data['editing_statuses']]
    extra_data['editing_status'].sort(key=lambda x: utils.normalized(x))
    extra_data['selling_status'] = [selling_status.name for selling_status
                                    in extra_data['selling_statuses']]
    extra_data['selling_status'].sort(key=lambda x: utils.normalized(x))
    extra_data['product_units'] = [product_unit.code for product_unit
                                   in extra_data['product_units']]
    extra_data['product_units'].sort(key=lambda x: utils.normalized(x))
    extra_data['tax_in'] = extra_data['tax_out'] = [0, 5, 10]

    return attributes_data, extra_data


def create_sample_import_file(attribute_set_id):
    """
    @Todo
    Create sample import file (excel) then return its url
    :param attribute_set_id:
    :return:
    """
    return None
    current_seller = {}

    manual_sku = False
    if current_seller:
        manual_sku = True
    s1_start_col = 30
    s2_start_col = 15
    bg_color = 'CCC0DA'

    attributes_data, extra_data = get_data_for_excel_sheet(attribute_set_id)

    # Get all attribute names of attribute set for specifications in SanPham sheet
    attributes = attributes_data.get('attributes', [])
    attribute_names = [attribute.name for attribute in attributes]

    # START - EXCEL process
    input_path = os.path.join(
        config.ROOT_DIR,
        'storage',
        'template',
        'template_create_product.xlsx'
    )
    output_dir = os.path.join(
        app.config['MEDIA_IMPORT_DIR'],
        'product',
        'template',
        str(current_seller.id) if manual_sku else 'default'
    )
    os.makedirs(output_dir, exist_ok=True)  # Make dir if not exist
    file_name = 'template_create_product_{}.xlsx'.format(
        utils.normalized(attributes_data['attribute_set_name']).replace(' ', '')
    )
    output_path = os.path.join(output_dir, file_name)

    book = load_workbook(input_path)
    ws1 = book['SanPham']
    ws2 = book['DuLieuMau']

    # Write attribute title to SanPham sheet
    if attribute_names:
        cell = ws1['{}7'.format(get_column_letter(s1_start_col))]
        cell.value = 'Thông số kỹ thuật'
        cell.font = Font(size='14', bold=True)
        cell.fill = PatternFill(start_color=bg_color, fill_type='solid')
        ws1.merge_cells('{}7:{}7'.format(get_column_letter(s1_start_col),
                                         get_column_letter(s1_start_col + len(attribute_names) - 1)))

    # Write attribute data to SanPham sheet
    for key, value in enumerate(attribute_names):
        cell = ws1['{}8'.format(get_column_letter(s1_start_col + key))]
        cell.value = value
        cell.fill = PatternFill(start_color=bg_color, fill_type='solid')

    if manual_sku:
        default_cols = [
            {'data': extra_data['types'], 's1_letter': 'B'},
            {'data': extra_data['categories'], 's1_letter': 'F'},
            {'data': extra_data['sale_categories'], 's1_letter': 'G'},
            {'data': extra_data['brands'], 's1_letter': 'H'},
            {'data': extra_data['colors'], 's1_letter': 'I'},
            {'data': extra_data['units'], 's1_letter': ['M', 'N']},
            {'data': extra_data['product_types'], 's1_letter': 'O'},
            {'data': extra_data['objectives'], 's1_letter': 'P'},
            {'data': extra_data['channels'], 's1_letter': 'Y'},
            {'data': extra_data['editing_status'], 's1_letter': ''},
            {'data': extra_data['selling_status'], 's1_letter': ''},
            {'data': extra_data['product_units'], 's1_letter': ''},
            {'data': extra_data['tax_in'], 's1_letter': 'W'},
            {'data': extra_data['tax_out'], 's1_letter': 'X'}
        ]
    else:
        # Write default data to DuLieuMau sheet
        default_cols = [
            {'data': extra_data['types'], 's1_letter': 'B'},
            {'data': extra_data['categories'], 's1_letter': 'E'},
            {'data': extra_data['sale_categories'], 's1_letter': 'F'},
            {'data': extra_data['brands'], 's1_letter': 'G'},
            {'data': extra_data['colors'], 's1_letter': 'H'},
            {'data': extra_data['units'], 's1_letter': ['L', 'M']},
            {'data': extra_data['product_types'], 's1_letter': 'N'},
            {'data': extra_data['objectives'], 's1_letter': 'O'},
            {'data': extra_data['channels'], 's1_letter': 'X'},
            {'data': extra_data['editing_status'], 's1_letter': ''},
            {'data': extra_data['selling_status'], 's1_letter': ''},
            {'data': extra_data['product_units'], 's1_letter': ''},
            {'data': extra_data['tax_in'], 's1_letter': 'V'},
            {'data': extra_data['tax_out'], 's1_letter': 'W'}
        ]
    for key, value in enumerate(default_cols):
        s2_letter = get_column_letter(key + 1)
        for v, c in zip(value['data'], ws2['{}2'.format(s2_letter):'{}{}'.format(s2_letter, len(value['data']) + 1)]):
            c[0].value = v
        # Add data validation dropdown list for SanPham sheet (with default data in DuLieuMau sheet)
        if value['s1_letter']:
            _add_data_validation_list(ws1, len(value['data']), s2_letter, value['s1_letter'])

    # Write attribute options to DuLieuMau sheet
    s2_attr_col_index = s2_start_col
    for attribute in attributes:
        if attribute.value_type in ['selection', 'multiple_select']:
            s2_letter = get_column_letter(s2_attr_col_index)
            # Header
            cell = ws2['{}1'.format(s2_letter)]
            cell.value = attribute.name
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=bg_color, fill_type='solid')
            # Content
            options = [option.value for option in attribute.options]
            for v, c in zip(options, ws2['{}2'.format(s2_letter):'{}{}'.format(s2_letter, len(options) + 1)]):
                c[0].value = v
            # Add data validation dropdown list for SanPham sheet (with default data in DuLieuMau sheet)
            if manual_sku:
                s1_letter = get_column_letter(s1_start_col + attribute_names.index(attribute.name) + 1)
            else:
                s1_letter = get_column_letter(s1_start_col + attribute_names.index(attribute.name))
            _add_data_validation_list(ws1, len(options), s2_letter, s1_letter)
            # Increase index
            s2_attr_col_index += 1

    if manual_sku:
        insert_sku_column(ws1)
    book.save(output_path)
    # END - EXCEL process

    template_url = "/".join(
        ['media', 'import', 'product', 'template', str(current_seller.id) if manual_sku else 'default', file_name]
    )

    return '{}{}'.format(request.host_url, template_url)


def insert_sku_column(ws1):
    ws1.insert_cols(3)
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 40
    sku_cell = ws1.cell(row=8, column=3)
    sku_cell.value = 'SKU'
    sku_cell.font = Font(bold=True, color='ff0000')
    sku_cell.fill = PatternFill(start_color='fdeada', fill_type='solid')
    for index_cell in range(2, 6):
        ws1.cell(row=index_cell, column=3).fill = PatternFill(start_color='dce6f2', fill_type='solid')
