# coding=utf-8

import os
from copy import copy
from operator import and_

from flask import send_file
from flask_login import current_user

from sqlalchemy.orm import (
    joinedload,
    load_only,
)
import openpyxl

import config
from catalog import models
from catalog.constants import UOM_CODE_ATTRIBUTE, IMPORT

from catalog.services.categories import category
from catalog.services.seller import get_default_platform_owner_of_seller


class TemplateBase:
    # Global variable here
    FILE_TEMPLATE = 'template_create_general_product_v2.0.xlsx'
    VAR_COL_OFFSET = 13
    TITLE_ROW_OFFSET = 5
    NUMBER_OF_DETAIL_INFO_COLUMNS = 9
    GROUP_TITLE_OFFSET = TITLE_ROW_OFFSET - 1
    COPY_FROM_CELL = 'C4'

    def __init__(self, import_type):
        self.n_cols_detail_info = self.NUMBER_OF_DETAIL_INFO_COLUMNS
        self.offset = self.VAR_COL_OFFSET
        self.file_path = os.path.join(config.ROOT_DIR, 'storage', 'template', self.FILE_TEMPLATE)
        self.wb = openpyxl.load_workbook(self.file_path)
        self.import_type = import_type

    def generate_general_product_template(self):
        pass

    def send_file(self):
        res = send_file(self.file_path, as_attachment=True)
        res.headers['Content-Disposition'] = f'attachment; filename="{os.path.basename(self.file_path)}"'
        return res

    # _______________________MANAGE COLUMN___________________________
    def _add_column_system_group_attributes(self, ws, attributes):
        # ================= Thuộc tính hệ thống ==================
        system_attributes = [x for x in attributes if bool(x.attribute_group.system_group)]
        if len(system_attributes) > 0:
            self.generate_group(
                ws, 'Thông số hệ thống',
                self.attribute_object_to_dict(
                    attributes=system_attributes,
                    force_required=False
                ),
                self.offset
            )
            self.offset += len(system_attributes)

    def _add_column_variation_attribute(self, ws, attributes):
        # ================= Thuộc tính biến thể ==================
        variation_attributes = [x for x in attributes if bool(x.is_variation)]
        if len(variation_attributes) > 0:
            self.generate_group(
                ws, 'Thuộc tính biến thể',
                self.attribute_object_to_dict(
                    attributes=variation_attributes,
                    force_required=True
                ),
                self.offset
            )
            self.offset += len(variation_attributes)

    def _add_column_not_variant_attribute(self, ws, attributes):
        # ================= Thông số kĩ thuật ================ ==
        normal_attributes = [x for x in attributes if not bool(x.is_variation)]
        if len(normal_attributes) > 0:
            self.generate_group(
                ws, 'Thông số kĩ thuật',
                self.attribute_object_to_dict(
                    attributes=normal_attributes,
                    force_required=False
                ),
                self.offset
            )
            self.offset += len(normal_attributes)

    def _add_column_variant_image(self, ws):
        # ================ Ảnh sản phẩm =======================
        image_cols = [{
            'name': 'Ảnh sản phẩm',
            'code': 'image urls',
            'description': 'Nhập danh sách url ảnh cách nhau bởi ký tự xuống dòng ( Alt + Enter)\nNhập tối đa 5 ảnh mỗi sản phẩm',
            'required': False,
        }]
        if len(image_cols) > 0:
            self.generate_group(ws, 'Ảnh sản phẩm', image_cols, self.offset)
            self.offset += len(image_cols)

    def _add_column_uom(self, ws):
        uom_cols = [{
            'name': 'Đơn vị tính',
            'code': 'uom',
            'description': 'Điền chính xác hoặc copy tên "Đơn vị tính" từ sheet DuLieuMau',
            'required': True,
        }, {
            'name': 'Tỷ lệ so với đơn vị tính gốc',
            'code': 'uom_ratio',
            'description': '"Nhập tỷ lệ quy đổi so với đơn vị gốc.\n- Nếu sản phẩm có nhiều đơn vị tính (ví dụ bán theo hộp, lốc, thùng...) thì bắt buộc nhập sản phẩm đơn vị gốc có tỷ lệ =1 trước, các sản phẩm còn lại sau.\n- Nếu sản phẩm chỉ có 1 đơn vị tính thì nhập 1"',
            'required': True
        }]
        self.generate_group(ws, 'Đơn vị tính', uom_cols, self.offset)
        self.offset += len(uom_cols)

    def _delete_column_seller_sku(self, ws):
        # ================= Thông tin chi tiết sản phẩm ==================
        ws.delete_cols(self.offset)
        self.n_cols_detail_info -= 1

    # _______________________MANAGE DATA___________________________
    def check_template_with_default_category(self):
        if self.import_type in IMPORT.IMPORT_WITH_DEFAULT_CATEGORY:
            return True

    def _load_category_data(self, title='Danh mục ngành hàng'):
        category_seller_id = current_user.seller_id
        if self.check_template_with_default_category():
            category_seller_id = get_default_platform_owner_of_seller(current_user.seller_id)
        categories = category.get_leaf_tree(
            seller_id=category_seller_id
        ).options(
            load_only('code', 'name')).all()
        if not title:
            return categories
        return {title: map(lambda x: f'{x.code}=>{x.name}', categories)}

    def _load_platform_category_data(self, platform_id, title='Danh mục ngành hàng'):
        platform_owner = models.PlatformSellers.query.filter(models.PlatformSellers.platform_id == platform_id,
                                                             models.PlatformSellers.is_owner.is_(True)).first()
        if not platform_owner:
            return {}
        categories = category.get_leaf_tree(
            seller_id=platform_owner.seller_id
        ).all()
        parent_ids = set()
        map_cats = {}
        for c in categories:
            path = list(map(lambda x: int(x), c.path.split('/')))
            map_cats[c.id] = (c, path)
            for id in path:
                parent_ids.add(id)
        parents = models.Category.query.filter(models.Category.id.in_(parent_ids)).all()
        parents = {i.id: i for i in parents}
        for (id, item) in map_cats.items():
            path = ''
            cat, cat_parent_ids = item
            for parent_id in cat_parent_ids:
                parent = parents.get(parent_id)
                if parent:
                    path = f'{path} / {parent.name}' if path else parent.name
            cat.ext_full_path_data = path

        return {title: list(map(lambda x: f'{x.id}=>{x.ext_full_path_data}', categories))}

    def _load_master_cateogry_data(self, title='Danh mục'):
        master_categories = category.get_leaf_tree(models.MasterCategory).options(
            load_only('code', 'name'))
        if not title:
            return master_categories
        return {title: map(lambda x: f'{x.code}=>{x.name}', master_categories)}

    def _load_brand_data(self, title='Thương hiệu'):
        brands = models.Brand.query.filter(
            models.Brand.is_active.is_(True)
        ).options(load_only('name')).order_by(models.Brand.name)
        if not title:
            return brands
        return {title: map(lambda x: x.name, brands)}

    def _load_unit_data(self, title='Đơn vị tính'):
        unit_attribute = models.Attribute.query.filter(
            models.Attribute.code == UOM_CODE_ATTRIBUTE
        ).first()  # return models.Attribute
        units = unit_attribute.select_options if unit_attribute else []

        if not title:
            return units
        return {title: map(lambda x: x.value, units)}

    def _load_product_type_data(self, title='Loại hình sản phẩm'):
        product_types = models.Misc.query.options(load_only('name')).filter(
            models.Misc.type == 'product_type').order_by(models.Misc.position)
        if not title:
            return product_types
        return {title: map(lambda x: x.name, product_types)}

    def _load_tax_data(self, title='Thuế suất'):
        taxes = models.Tax.query.options(load_only('label')).order_by(models.Tax.id)
        if not title:
            return taxes
        return {title: map(lambda x: x.label, taxes)}

    def _load_shipping_type_data(self, title='Loại hình vận chuyển'):
        shipping_types = models.ShippingType.query.filter(
            models.ShippingType.is_active.is_(True)
        ).options(load_only('name')).order_by(models.ShippingType.name).all()
        if not title:
            return shipping_types
        return {title: map(lambda x: x.name, shipping_types)}

    def _load_attribute_sets_data(self, title='Nhóm sản phẩm', with_variant=True):
        if with_variant:
            attribute_sets = models.AttributeSet.query.options(load_only('id', 'name'))
        else:
            has_variant_attribute_set_ids = models.AttributeSet.query.join(
                models.AttributeGroup,
                and_(
                    models.AttributeGroup.attribute_set_id == models.AttributeSet.id,
                    models.AttributeGroup.system_group != 1
                )
            ).join(
                models.AttributeGroupAttribute,
                models.AttributeGroupAttribute.attribute_group_id == models.AttributeGroup.id
            ).filter(
                models.AttributeGroupAttribute.is_variation == 1
            ).options(load_only('id')).all()

            attribute_sets = models.AttributeSet.query.filter(
                models.AttributeSet.id.notin_([attribute_set.id for attribute_set in has_variant_attribute_set_ids])
            ).options(load_only('id', 'name'))

        if not title:
            return attribute_sets
        return {title: map(lambda x: f'{x.id}=>{x.name}', attribute_sets)}

    def _load_sample_data_queries(self):
        sample_data_queries = {
            **self._load_category_data(),
            **self._load_master_cateogry_data(),
            **self._load_brand_data(),
            **self._load_unit_data(),
            **self._load_product_type_data(),
            **self._load_tax_data(),
            **self._load_shipping_type_data()
        }
        return sample_data_queries

    def _load_sample_data_attribute_set(self, attribute_set_id):
        attribute_set = models.AttributeSet.query.get(attribute_set_id)
        return attribute_set

    def _load_sample_data_attributes(self, attribute_set_id):
        attributes = models.db.session.query(models.AttributeGroupAttribute).join(
            models.AttributeGroup,
            models.AttributeGroupAttribute.attribute_group_id == models.AttributeGroup.id
        ).join(
            models.Attribute,
            models.Attribute.id == models.AttributeGroupAttribute.attribute_id
        ).filter(
            models.Attribute.code.notin_(['uom', 'uom_ratio', ]),
            models.AttributeGroup.attribute_set_id == attribute_set_id
        ).options(
            load_only('is_variation'),
            joinedload('attribute_group').load_only('system_group'),
            joinedload('attribute').load_only('name', 'display_name', 'code', 'value_type').options(
                joinedload('options').load_only('value')
            )
        ).order_by(models.AttributeGroup.priority, models.AttributeGroupAttribute.priority).all()
        return attributes

    def _load_sample_data_attribute_options(self, attributes):
        sample_data_queries = {}
        for attr_info in attributes:
            if attr_info.attribute.value_type in ('selection', 'multiple_select'):
                option_key = attr_info.attribute.display_name or attr_info.attribute.name
                sample_data_queries[option_key] = map(
                    lambda x: x.value,
                    attr_info.attribute.options
                )
        return sample_data_queries

    # _______________________WORK WITH EXCEL FILE__________________________
    def _generate_sample_data_sheet(self, ws, queries, curr_col=2):
        title_row = 1
        for field_name, items in queries.items():
            ws.cell(row=title_row, column=curr_col, value=field_name)
            curr_row = title_row + 1
            for item in items:
                if item.startswith('='):
                    item = '\'' + item
                ws.cell(row=curr_row, column=curr_col, value=item)
                curr_row += 1
            curr_col += 1
        title_row = ws.row_dimensions[1]
        title_row.font = openpyxl.styles.Font(bold=True)

    def _copy_style(self, dest, src, fields):
        for field in fields:
            if hasattr(src, field):
                setattr(dest, field, copy(getattr(src, field)))

    def _merge_cells(self, ws, start_column, end_column, copy_cell_at, cell_name=None):
        ws.merge_cells(
            start_row=self.GROUP_TITLE_OFFSET,
            end_row=self.GROUP_TITLE_OFFSET,
            start_column=start_column,
            end_column=end_column
        )
        cell = ws[self.GROUP_TITLE_OFFSET][start_column - 1]
        if cell_name:
            cell.value = cell_name
        self._copy_style(cell, ws[copy_cell_at], ('fill', 'font', 'alignment'))

    def generate_group(self, ws, name_group, items, offset):
        n_items = len(items)
        ws.insert_cols(offset, n_items)
        for idx, item in enumerate(items):
            name_cell = ws.cell(self.TITLE_ROW_OFFSET, offset + idx, item.get('display_name') or item['name'])
            self.format_name_cell(name_cell, item['required'])

            code_cell = ws.cell(self.TITLE_ROW_OFFSET + 1, offset + idx, item['code'])
            self.format_code_cell(code_cell)

            description_cell = ws.cell(self.TITLE_ROW_OFFSET + 2, offset + idx, item['description'])
            self.format_description_cell(description_cell)

        ws.merge_cells(
            start_row=self.GROUP_TITLE_OFFSET,
            end_row=self.GROUP_TITLE_OFFSET,
            start_column=offset,
            end_column=offset + n_items - 1
        )
        cell = ws[self.GROUP_TITLE_OFFSET][offset - 1]
        cell.value = name_group
        self._copy_style(cell, ws[self.COPY_FROM_CELL], ('fill', 'font', 'alignment'))

    def format_name_cell(self, cell, required=False):
        cell.fill = openpyxl.styles.PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid')
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='bottom')
        no_border = openpyxl.styles.Side(style=None)
        cell.border = openpyxl.styles.Border(no_border, no_border, no_border, no_border)
        if required:
            cell.font = openpyxl.styles.Font(size=11, name='Times New Roman', color='FF0000', bold=True)
        else:
            cell.font = openpyxl.styles.Font(size=11, name='Times New Roman', color='000000', bold=True)

    def format_code_cell(self, cell):
        cell.font = openpyxl.styles.Font(size=10, name='Times New Roman', color='808080', bold=False)
        cell.fill = openpyxl.styles.PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid')
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='bottom')
        no_border = openpyxl.styles.Side(style=None)
        cell.border = openpyxl.styles.Border(no_border, no_border, no_border, no_border)

    def format_description_cell(self, cell):
        cell.font = openpyxl.styles.Font(size=11, name='Times New Roman', color='000000', bold=False)
        cell.fill = openpyxl.styles.PatternFill(start_color='F3F4F3', end_color='F3F4F3', fill_type='solid')
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        side = openpyxl.styles.Side(color='CFCFCF', border_style='thin')
        cell.border = openpyxl.styles.Border(side, side, side, side)

    def attribute_object_to_dict(self, attributes, description=None, force_required=False):
        """
        :param: attributes, list[AttributeGroupAttribute]
        """
        ret = []
        for x in attributes:
            ret.append({
                'name': x.attribute.name,
                'display_name': x.attribute.display_name,
                'code': x.attribute.code,
                'description': x.attribute.description,
                'required': force_required,
            })
        return ret
