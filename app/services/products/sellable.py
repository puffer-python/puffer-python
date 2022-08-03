# coding=utf-8

import io
import logging
import os
from typing import List
import funcy
import openpyxl
import sqlalchemy

from time import strftime
from collections import defaultdict
from sqlalchemy.orm import (
    joinedload,
    load_only,
    noload,
)
from sqlalchemy.sql.expression import cast, text, exists
from flask_login import current_user
from funcy import lpluck_attr
from sqlalchemy import or_, and_, case, func

from catalog import models as m, utils
from catalog.constants import UOM_CODE_ATTRIBUTE, ExportSellable, MAX_RECORD
from catalog.extensions.exceptions import BadRequestException
from catalog.models import db
from catalog.extensions import (
    exceptions as exc,
    signals
)
from catalog.models.sellable_product import SellableProduct
from catalog.services import QueryBase
from catalog.services import shipping_policy as svr
from catalog.services import seller as seller_services
from catalog.services.attribute_sets.attribute_set import get_variant_attribute_by_attribute_set_id
from catalog.services.attributes import AttributeService
from catalog.services.shipping_types.sellable_product_shipping_type import SellableProductShippingTypeService
from catalog.services.shipping_types.shipping_type import get_default_shipping_type
from catalog.utils import (
    cast_separated_string_to_ints,
    safe_cast,
)
from catalog.utils.lambda_list import LambdaList
from catalog.validators import sellable as sellable_validator
from config import ROOT_DIR
from catalog.models.db_constants import AttributeValueType
from catalog.constants import DIMENSION_ATTRIBUTES_CODES, PACK_CODE_ATTRIBUTES

__author__ = 'Kien.HT'

_logger = logging.getLogger(__name__)


def save_changes():
    m.db.session.commit()


def _get_same_uom(sellable):
    return m.SellableProduct.query.filter(
        m.SellableProduct.uom_code == sellable.uom_code,
        m.SellableProduct.product_id == sellable.product_id,
        m.SellableProduct.seller_id == sellable.seller_id,
        m.SellableProduct.uom_ratio != sellable.uom_ratio
    ).first()


def gen_new_sku():
    """

    :return:
    """
    now = strftime('%y%m')

    _like_expr = '{}%'.format(now)
    _number_expr = '^[0-9]+$'

    MIN_SKU_LENGTH = 9
    MAX_SKU_LENGTH = 10

    max_sku = m.db.session.query(func.max(m.SellableProduct.sku)).filter(
        m.SellableProduct.sku.like(_like_expr),
        # check if sku is numeric - cannot use regexp because of sqlite
        m.SellableProduct.sku == cast(cast(m.SellableProduct.sku, sqlalchemy.Integer), sqlalchemy.String),
        func.length(m.SellableProduct.sku) >= MIN_SKU_LENGTH,
        func.length(m.SellableProduct.sku) <= MAX_SKU_LENGTH,
        m.SellableProduct.is_bundle == 0
    )
    max_sku = max_sku.first()[0]

    if max_sku:
        try:
            return str(int(max_sku) + 1)
        except ValueError as ex:
            raise ex

    return str(int(now) * 100000)


def gen_new_bundle_sku():
    """

    :return:
    """
    now = strftime('%y%m')

    _like_expr = 'B{}%'.format(now)

    max_sku = m.SellableProduct.query.filter(
        m.SellableProduct.sku.like(_like_expr),
        func.length(m.SellableProduct.sku) >= 9,
        m.SellableProduct.is_bundle == 1
    ).order_by(m.SellableProduct.sku.desc()).first()

    if max_sku:
        return 'B{}'.format(int(max_sku.sku[1:]) + 1)

    return 'B{}'.format(int(now) * 10000)


def create_sellable_products(data, __not_bulk_commit=False, seller=None, sellable_create_signal=True):
    sellable_products = []
    product = m.Product.query.get(data['product_id'])
    if seller is None:
        seller = seller_services.get_seller_by_id(current_user.seller_id)
    for product_data in data['sellable_products']:
        sellable_product = create_sellable_product(
            product_data=product_data,
            product=product,
            seller=seller,
            autocommit=False
        )
        sellable_products.append(sellable_product)

    m.db.session.flush()
    # update product status
    product.editing_status_code = 'processing'

    if not __not_bulk_commit:
        m.db.session.commit()

    if sellable_create_signal:
        # TODO: check if need publish event AddVariantSkuMsg to Clearance svc
        for item in sellable_products:
            signals.sellable_create_signal.send(item)

    return sellable_products, 'Tạo SKU thành công'


def _find_diff_attribute(from_attribute_set_id, to_attribute_set_id):
    from_attributes = get_variant_attribute_by_attribute_set_id(from_attribute_set_id)
    if not from_attributes:
        return None
    to_attributes = get_variant_attribute_by_attribute_set_id(to_attribute_set_id)
    if not to_attributes:
        return None
    for to_attribute in to_attributes:
        if to_attribute.id not in funcy.lpluck_attr('id', from_attributes):
            raise BadRequestException(
                'Danh mục mới chứa thuộc tính {} biến thể không đồng nhất'.format(to_attribute.name))

    return from_attributes, to_attributes


def _update_category_sellable_product(sellable, category_id, created_by):
    if sellable.category_id == category_id:
        return None
    sellable.category_id = category_id
    update_category = m.Category.query.filter(m.Category.id == category_id).first()
    product_category = m.ProductCategory.query.filter(
        m.ProductCategory.product_id == sellable.product_id,
        m.Category.id == m.ProductCategory.category_id,
        m.Category.seller_id == update_category.seller_id
    ).first()
    if product_category:
        product_category.category_id = category_id
    else:
        product_category = m.ProductCategory(
            product_id=sellable.product_id,
            category_id=category_id,
            created_by=created_by
        )
        m.db.session.add(product_category)
    if not update_category.default_attribute_set or not sellable.category.default_attribute_set:
        raise BadRequestException('Không thể cập nhật sang Danh mục không có Bộ thuộc tính')
    if sellable.category.default_attribute_set.id != update_category.default_attribute_set.id:
        _find_diff_attribute(sellable.category.default_attribute_set.id, update_category.default_attribute_set.id)
        sellable.attribute_set_id = update_category.default_attribute_set.id
        sellable.product.attribute_set_id = update_category.default_attribute_set.id


def _update_model_sellable_product(sellable, model):
    m.Product.query.get(sellable.product_id).model = model
    m.SellableProduct.query.filter(
        m.SellableProduct.product_id == sellable.product_id
    ).update({'model': model}, synchronize_session=False)


def _insert_sellable_product_barcodes(sellable_product_id, barcodes_with_source, created_by, update=False):
    if barcodes_with_source:
        if update:
            m.db.session.query(m.SellableProductBarcode).filter(
                m.SellableProductBarcode.sellable_product_id == sellable_product_id
            ).update({'is_default': False})

            # Remove product barcodes before updating the product itself
            m.db.session.query(m.SellableProductBarcode).filter(
                m.SellableProductBarcode.sellable_product_id == sellable_product_id,
                m.SellableProductBarcode.barcode.in_(funcy.lpluck('barcode', barcodes_with_source))
            ).delete(synchronize_session='fetch')
        for item in barcodes_with_source:
            is_default = (item.get('barcode') == barcodes_with_source[-1].get('barcode'))
            sku_barcode = m.SellableProductBarcode(
                sellable_product_id=sellable_product_id,
                barcode=item.get('barcode'),
                source=item.get('source'),
                created_by=created_by,
                is_default=is_default
            )
            m.db.session.add(sku_barcode)


def _delete_sellable_product_barcodes(sellable_product_id):
    m.db.session.query(m.SellableProductBarcode).filter(
        m.SellableProductBarcode.sellable_product_id == sellable_product_id
    ).delete()


def create_sellable_product(product_data, product, seller, autocommit=True):
    """
    :param product_data:
    :param product:
    :param seller: dict
    :param autocommit:
    :param created_by:
    :return:
    """

    if 'created_by' not in product_data:
        product_data['created_by'] = current_user.email

    variant = m.ProductVariant.query.get(product_data.get('variant_id'))
    sellable = m.SellableProduct()
    sellable.variant_id = product_data.get('variant_id')
    sellable.name = product_data.get('name') or variant.name
    sellable.supplier_sale_price = product_data.get('supplier_sale_price')
    sellable.allow_selling_without_stock = product_data.get(
        'allow_selling_without_stock', False)
    sellable.expiry_tracking = product_data.get('expiry_tracking')
    sellable.expiration_type = product_data.get('expiration_type')
    sellable.tracking_type = product_data.get('tracking_type', False) or product_data.get('manage_serial', False)
    sellable.days_before_exp_lock = product_data.get('days_before_exp_lock')
    sellable.provider_id = product_data.get('provider_id', seller.get('id'))
    sellable.product_id = product.id
    sellable.seller_id = seller.get('id')
    sellable.brand_id = product.brand_id
    sellable.category_id = product.category_id
    sellable.master_category_id = product.master_category_id
    sellable.attribute_set_id = product.attribute_set_id
    sellable.model = product.model
    sellable.warranty_months = product.warranty_months
    sellable.warranty_note = product.warranty_note
    sellable.tax_in_code = product.tax_in_code
    sellable.tax_out_code = product.tax_out_code
    sellable.product_type = product_data.get('product_type') or product.type
    sellable.master_category_id = product.master_category_id

    #  get unit from variant if available, else fallback to product
    sellable.unit_id = getattr(variant.unit, 'id', product.unit_id)
    sellable.unit_po_id = getattr(variant.unit, 'id', product.unit_id)
    sellable.uom_code = variant.unit.code
    sellable.uom_ratio = variant.uom_ratio
    sellable.uom_name = variant.unit.value
    sellable.updated_by = product_data.get('created_by')
    barcodes_with_source = product_data.get('barcodes')
    if not product.is_bundle:
        sellable.barcode = product_data.get('barcode')
        if barcodes_with_source:
            sellable.barcode = barcodes_with_source[-1].get('barcode')
        sellable.part_number = product_data.get('part_number')
        sellable.manage_serial = product_data.get('manage_serial')
        sellable.auto_generate_serial = product_data.get('auto_generate_serial')
        sellable.sku = gen_new_sku()
        if not seller.get('isAutoGeneratedSKU'):
            sellable.seller_sku = product_data.get('seller_sku', product_data.get('sku', sellable.sku))
            # if seller.get('code') not in SELLERS_WITH_MULTI_UOM:
            #     sellable.sku = sellable.seller_sku
        else:
            if float(sellable.uom_ratio) == 1:
                sellable.seller_sku = sellable.sku
            else:
                base_uom_id = int(variant.base_uom_id)
                base_uom_sku = m.db.session.query(m.SellableProduct). \
                    filter(m.SellableProduct.variant_id == base_uom_id).first()
                if base_uom_sku:
                    sellable.seller_sku = base_uom_sku.seller_sku

    else:
        sellable.sku = gen_new_bundle_sku()
        sellable.seller_sku = sellable.sku
        sellable.is_bundle = True

    base_same_uom = _get_same_uom(sellable)
    if base_same_uom:
        sellable.need_convert_qty = 1
        base_same_uom.need_convert_qty = 1
    m.db.session.add(sellable)
    if seller.get('servicePackage') == 'FBS':
        sellable.selling_status_code = 'hang_ban'
    sellable.editing_status_code = 'processing'
    m.db.session.flush()

    seo_terminal = m.SellableProductSeoInfoTerminal(
        terminal_id=0,
        sellable_product_id=sellable.id,
        short_description=product_data.get('short_description') or product.description,
        description=product_data.get('description') or product.detailed_description,
        created_by=product_data.get('created_by'),
        updated_by=product_data.get('created_by'),
    )
    m.db.session.add(seo_terminal)
    m.db.session.flush()

    shipping_type_ids = product_data.get('shipping_types')
    if shipping_type_ids:

        for shipping_type_id in shipping_type_ids:
            SellableProductShippingTypeService.create(
                sellable_product_id=sellable.id,
                shipping_type_id=shipping_type_id,
                auto_commit=False
            )
    else:
        default_shipping_type = get_default_shipping_type()
        if default_shipping_type:
            SellableProductShippingTypeService.create(
                sellable_product_id=sellable.id,
                shipping_type_id=default_shipping_type.id,
                auto_commit=False
            )
        m.db.session.flush()

    if not product.is_bundle:
        _insert_sellable_product_barcodes(sellable.id, barcodes_with_source, product_data.get('created_by'))

    if autocommit:
        m.db.session.commit()
    return sellable


def get_terminals_of_sellable_product(sellable_id):
    """

    :param sellable_id:
    :return:
    """

    def _get_terminal_codes_from_dict(terminal_dict):
        codes = terminal_dict['terminal_codes']
        if all([code is None for code in codes]):
            return 'all'

        return codes

    query = m.SellableProductTerminal.query
    query = query.outerjoin(m.Terminal)
    query = query.filter(
        m.SellableProductTerminal.sellable_product_id == sellable_id
    )
    results = query.all()  # type: list[m.SellableProductTerminal]

    terminal_data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for item in results:
        terminal_data[item.apply_seller_id][item.terminal_type]['terminal_codes'].append(
            item.terminal_code
        )

    return [{
        'apply_seller_id': seller_id,
        'terminals': [{
            'terminal_type': terminal_type,
            'terminal_codes': _get_terminal_codes_from_dict(terminal_codes)
        } for terminal_type, terminal_codes in seller_terminal_data.items()]
    } for seller_id, seller_terminal_data in terminal_data.items()]


def get_common_sellable_product_detail(sellable_product):
    shipping_property = svr.get_shipping_property_of_sellable_product(
        sellable_product)
    setattr(sellable_product, 'shipping_property', shipping_property.upper())
    shipping_types = db.session.query(
        m.ShippingType
    ).join(
        m.SellableProductShippingType,
        m.SellableProductShippingType.shipping_type_id == m.ShippingType.id
    ).filter(m.SellableProductShippingType.sellable_product_id == sellable_product.id).all()
    setattr(sellable_product, 'shipping_types', shipping_types)
    return sellable_product


def get_uoms_of_sellable_product(sellable):
    """
    Get all UoM SKUs by querying SKU -> variant -> all_uom_ratios -> variant_ids -> SKUs.
    These SKUs must belong to the same seller.

    :param sellable:
    :return:
    """
    uom_variant_ids = sellable.product_variant.extract_uom_ids()
    uom_skus = m.SellableProduct.query.filter(
        m.SellableProduct.variant_id.in_(uom_variant_ids),
        m.SellableProduct.seller_id == current_user.seller_id
    ).all()

    return {f'{uom_sku.sku}': {
        'uomCode': uom_sku.uom_code,
        'uomRatio': uom_sku.uom_ratio,
        'uomName': uom_sku.uom_name
    } for uom_sku in uom_skus}


def get_sellable_product_detail(product_id, data_key):
    """

    :param product_id:
    :param data_key:
    :return:
    """
    sellable = m.SellableProduct.query.get(product_id)  # type: m.SellableProduct
    if not sellable:
        raise exc.NotFoundException()

    if sellable.seller_id != current_user.seller_id:
        raise exc.NotFoundException()

    variant = sellable.product_variant

    data_map = {
        'common': get_common_sellable_product_detail(sellable),
        'images': variant.images,
        'specs': variant.variant_attributes,
        'terminals': get_terminals_of_sellable_product(sellable.id),
        'terminal_groups': get_terminal_groups_of_sellable_product(sellable.id),
        'uom': get_uoms_of_sellable_product(sellable)
    }
    if data_key not in data_map:
        raise exc.BadRequestException('Invalid data key')
    return {data_key: data_map[data_key]}


def get_sellable_product_detail_by_sku(sku, data_key):
    """

    :param product_id:
    :param data_key:
    :return:
    """
    sellable = m.SellableProduct.query.filter(m.SellableProduct.sku == sku).first()  # type: m.SellableProduct
    if not sellable:
        raise exc.NotFoundException()

    if sellable.seller_id != current_user.seller_id:
        raise exc.NotFoundException()

    variant = sellable.product_variant

    data_map = {
        'common': get_common_sellable_product_detail(sellable),
        'images': variant.images,
        'specs': variant.variant_attributes,
        'terminals': get_terminals_of_sellable_product(sellable.id),
        'terminal_groups': get_terminal_groups_of_sellable_product(sellable.id),
        'uom': get_uoms_of_sellable_product(sellable)
    }
    if data_key not in data_map:
        raise exc.BadRequestException('Invalid data key')
    return {data_key: data_map[data_key]}


def get_by_sku(sku: str, only_fields: tuple):
    # TODO: check why cannot load only variant_id field with load_only
    # error: Can't find property named "v" on mapped class SellableProduct->sellable_products in this Query
    sellable = m.SellableProduct.query.filter(
        m.SellableProduct.sku == sku,
    ).first()

    return sellable


def set_sellable_terminal(skus, seller_terminals):
    """

    :param skus:
    :param seller_terminals:
    :return:
    """
    sellable_products = m.SellableProduct.query.filter(
        m.SellableProduct.sku.in_(skus)
    ).all()

    changed = False
    for sellable in sellable_products:
        terminal_data = get_terminals_of_sellable_product(sellable.id)
        if terminal_data != seller_terminals:
            changed = True

    if not changed:
        return {
            'skus': skus,
            'seller_terminals': seller_terminals
        }

    # delete old records
    m.SellableProductTerminal.query.filter(
        m.SellableProductTerminal.sellable_product_id.in_(
            lpluck_attr('id', sellable_products)
        )
    ).delete(False)

    for terminal_data in seller_terminals:
        apply_seller_id = terminal_data.get('apply_seller_id')
        set_seller_terminals(
            seller_id=apply_seller_id,
            terminal_list=terminal_data.get('terminals'),
            sellable_products=sellable_products
        )

    m.db.session.commit()

    for sellable in sellable_products:
        signals.sellable_update_signal.send(sellable)

    return {
        'skus': skus,
        'seller_terminals': seller_terminals,
    }


def set_seller_terminals(seller_id, terminal_list, sellable_products):
    """

    :param seller_id:
    :param terminal_list:
    :param sellable_products:
    :return:
    """
    for terminal_data in terminal_list:
        terminal_codes = terminal_data.get('terminal_codes')
        terminals = m.Terminal.query.filter(
            m.Terminal.code.in_(terminal_codes)
        ).all() if terminal_codes != 'all' else [0]
        for terminal in terminals:
            for sellable in sellable_products:
                res = m.SellableProductTerminal()
                res.terminal_id = getattr(terminal, 'id', 0)
                res.terminal_code = getattr(terminal, 'code', None)
                res.sellable_product_id = sellable.id
                res.terminal_type = terminal_data.get('terminal_type')
                res.apply_seller_id = seller_id
                res.on_off_status = 'on'
                res.created_by = current_user.email
                m.db.session.add(res)

    m.db.session.flush()


class SellableProductListQuery(QueryBase):
    model = m.SellableProduct

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.not_found = None
        self.query = self.query.order_by(
            m.SellableProduct.product_id.desc()
        )

    def apply_filters(self, restrict_seller=True, **kwargs):
        seller = kwargs.get('seller') # TODO: check APIs accept "seller" param
                                      # as comma-separeted Seller IDs string to remove this
        if kwargs.get('seller_id') != 0:
            if restrict_seller:
                search_seller_id = kwargs.get('seller_id') or current_user.seller_id
                search_seller_id = safe_cast(search_seller_id, int)
                if search_seller_id:
                    self._apply_seller_filter([search_seller_id])
            elif seller:
                self._apply_seller_terminal_filter(
                    [int(sid) for sid in seller.split(',')]
                )

        seller_ids = kwargs.get('seller_ids')
        seller_id = None
        if seller_ids is not None:
            if not seller_ids:
                self.obvious_not_found = True
                return

            seller_id = seller_ids[0]
            self._apply_seller_ids_filter(seller_ids)

        product_ids = kwargs.get('product_ids')
        if product_ids is not None:
            if not product_ids:
                self.obvious_not_found = True
                return

            self._apply_product_ids_filter(product_ids)

        variant_ids = kwargs.get('variant_ids')
        if variant_ids is not None:
            if not variant_ids:
                self.obvious_not_found = True
                return

            self._apply_variant_ids_filter(variant_ids)

        ids = kwargs.get('ids')
        if ids:
            self._apply_id_filter(ids)

        skus = kwargs.get('skus')
        if skus:
            self._apply_skus_filter(skus)

        seller_skus = kwargs.get('seller_skus')
        if seller_skus:
            self._apply_seller_skus_filter(seller_skus)

        barcodes = kwargs.get('barcodes')
        if barcodes:
            self._apply_barcodes_filter(barcodes)

        restrict_convert_qty = kwargs.get('restrict_convert_qty')
        if restrict_convert_qty:
            self._apply_restrict_convert_qty_filter()

        keyword = kwargs.get('keyword')
        if keyword:
            self._apply_keyword_filter(keyword)

        provider_ids = kwargs.get('provider_ids')
        if provider_ids is not None:
            if not provider_ids:
                self.obvious_not_found = True
                return

            self._apply_provider_filter(provider_ids)

        category_ids_str = kwargs.get('category')
        if category_ids_str:
            self._apply_category_filter(
                cast_separated_string_to_ints(category_ids_str),
                restrict_seller,
            )

        category_ids_str = kwargs.get('category_ids')
        if category_ids_str:
            self._apply_category_filter(
                cast_separated_string_to_ints(category_ids_str),
                restrict_seller,
                seller_id=seller_id
            )

        master_cat_ids_str = kwargs.get('master_category') or kwargs.get('master_category_ids')
        if master_cat_ids_str:
            self._apply_master_category_filter(
                cast_separated_string_to_ints(master_cat_ids_str)
            )

        brand_ids_str = kwargs.get('brand') or kwargs.get('brand_ids')
        if brand_ids_str:
            self._apply_brand_filter(
                cast_separated_string_to_ints(brand_ids_str)
            )

        attr_set_ids_str = kwargs.get('attribute_set') or kwargs.get('attribute_set_ids')
        if attr_set_ids_str:
            self._apply_attribute_set_filter(
                cast_separated_string_to_ints(attr_set_ids_str)
            )

        selling_status = kwargs.get('selling_status')
        if selling_status:
            self._apply_selling_status_filter(
                status_ids=selling_status
            )

        editing_status = kwargs.get('editing_status')
        if editing_status:
            self._apply_editing_status_filter(
                status_ids=editing_status
            )

        editing_status_codes = kwargs.get('editing_status_codes')
        if editing_status_codes:
            self._apply_editing_status_codes_filter(editing_status_codes)

        terminal = kwargs.get('terminal')
        if terminal:
            self._apply_terminal_filter(
                terminal_code=terminal,
                on_off_status=kwargs.get('on_off_status')
            )

        terminal_group = kwargs.get('terminal_group')
        if terminal_group:
            self._apply_terminal_group_filter(
                terminal_group_code=terminal_group
            )

        is_bundle = kwargs.get('is_bundle')
        if is_bundle is not None:
            self._apply_bundle_filter(is_bundle)

        models = kwargs.get('models')
        if models is not None:
            self._apply_group_model_filter(models=models)

    def _apply_seller_ids_filter(self, seller_ids):
        self.query = self.query.filter(
            m.SellableProduct.seller_id.in_(seller_ids)
        )

    def _apply_product_ids_filter(self, product_ids):
        self.query = self.query.filter(
            m.SellableProduct.product_id.in_(product_ids)
        )

    def _apply_variant_ids_filter(self, variant_ids):
        self.query = self.query.filter(
            m.SellableProduct.variant_id.in_(variant_ids)
        )

    def _apply_id_filter(self, ids):
        self.query = self.query.filter(
            m.SellableProduct.id.in_(ids)
        )

    def _apply_skus_filter(self, skus):
        self.query = self.query.filter(
            m.SellableProduct.sku.in_(skus)
        )

    def _apply_seller_skus_filter(self, seller_skus):
        self.query = self.query.filter(
            m.SellableProduct.seller_sku.in_(seller_skus)
        )

    def _apply_barcodes_filter(self, barcodes):
        alias_barcodes = m.db.aliased(m.SellableProductBarcode)
        self.query = self.query.filter(
            alias_barcodes.query.filter(
                alias_barcodes.sellable_product_id == m.SellableProduct.id,
                alias_barcodes.barcode.in_(barcodes)
            ).exists()
        )

    def _apply_seller_filter(self, seller_ids):
        self.query = self.query.filter(
            m.SellableProduct.seller_id.in_(seller_ids)
        )

    def _apply_provider_filter(self, provider_ids):
        self.query = self.query.filter(
            m.SellableProduct.provider_id.in_(provider_ids)
        )

    def _apply_seller_terminal_filter(self, seller_ids):
        self.query = self.query.filter(
            m.SellableProductTerminal.apply_seller_id.in_(seller_ids)
        )

    def _apply_keyword_filter(self, kw):
        name = ''.join(kw)
        alias_barcodes = m.db.aliased(m.SellableProductBarcode)
        self.query = self.query.filter(
            or_(
                m.SellableProduct.name.ilike(f'%{name}%'),
                alias_barcodes.query.filter(
                    alias_barcodes.sellable_product_id == m.SellableProduct.id,
                    alias_barcodes.barcode.ilike(f'%{name}%')
                ).exists(),
                m.SellableProduct.seller_sku.in_(kw),
                m.SellableProduct.sku.in_(kw)
            )
        )

    def _apply_category_filter(self, category_ids: List[int], restrict_seller, seller_id=None):
        def _filter_by_category_ids(list_category_ids):
            if list_category_ids:
                self.query = self.query.filter(
                    m.ProductCategory.query.filter(
                        m.ProductCategory.product_id == m.SellableProduct.product_id,
                        m.ProductCategory.category_id.in_(list_category_ids)
                    ).exists())

        query = m.Category.query.filter(
            m.Category.id.in_(category_ids),
            m.Category.is_active.is_(True)
        )
        if restrict_seller:
            query = query.filter(
                m.Category.seller_id == current_user.seller_id
            )
        categories = query.all()

        if not categories:
            self.query = self.query.filter(False)
            return

        paths = [category.path for category in categories]
        # select all categories, include children
        full_categories = m.Category.query.filter(
            or_(
                *[m.Category.path.like('{}'.format(path)) for path in paths],
                *[m.Category.path.like('{}/%'.format(path)) for path in paths]
            ),
            m.Category.is_active.is_(True)
        ).all() if paths else []
        cat_ids = []
        platform_cat_ids = []
        seller_id = safe_cast(seller_id, int)
        if seller_id:
            owner_seller_id = seller_services.get_default_platform_owner_of_seller(seller_id)
            for c in full_categories:
                if c.seller_id == owner_seller_id:
                    cat_ids.append(c.id)
                else:
                    platform_cat_ids.append(c.id)
        else:
            cat_ids = [category.id for category in full_categories]
        _filter_by_category_ids(cat_ids)
        _filter_by_category_ids(platform_cat_ids)

    def _apply_master_category_filter(self, category_ids: List[int]):
        query = m.MasterCategory.query.filter(
            m.MasterCategory.id.in_(category_ids),
            m.MasterCategory.is_active.is_(True)
        )
        categories = query.all()

        paths = [category.path for category in categories]
        # select all categories, include children
        full_categories = m.MasterCategory.query.filter(
            or_(
                *[m.MasterCategory.path.like('{}'.format(path)) for path in paths],
                *[m.MasterCategory.path.like('{}/%'.format(path)) for path in paths]
            ),
            m.MasterCategory.is_active.is_(True)
        ).all() if paths else []
        cat_ids = [category.id for category in full_categories]
        self.query = self.query.filter(
            m.SellableProduct.master_category_id.in_(cat_ids)
        )

    def _apply_brand_filter(self, brand_ids: List[int]):
        self.query = self.query.join(m.Brand)
        self.query = self.query.filter(
            m.SellableProduct.brand_id.in_(brand_ids),
            m.Brand.is_active.is_(True)
        )

    def _apply_attribute_set_filter(self, attribute_set_ids: List[int]):
        self.query = self.query.filter(
            m.SellableProduct.attribute_set_id.in_(attribute_set_ids)
        )

    def _apply_editing_status_filter(self, status_ids):
        statuses = m.EditingStatus.query.filter(
            m.EditingStatus.id.in_(status_ids.split(','))
        ).all()
        status_codes = [status.code for status in statuses]
        self.query = self.query.filter(
            m.SellableProduct.editing_status_code.in_(status_codes)
        )
        self.query = self.query

    def _apply_editing_status_codes_filter(self, editing_status_codes):
        self.query = self.query.filter(
            m.SellableProduct.editing_status_code.in_(editing_status_codes)
        )
        self.query = self.query

    def _apply_selling_status_filter(self, status_ids):
        statuses = m.SellingStatus.query.filter(
            m.SellingStatus.id.in_(status_ids.split(','))
        ).all()
        status_codes = [status.code for status in statuses]
        self.query = self.query.filter(
            m.SellableProduct.selling_status_code.in_(status_codes)
        )

    def _apply_terminal_filter(self, terminal_code, on_off_status):
        terminal = m.Terminal.query.filter(
            m.Terminal.code == terminal_code,
        ).options(
            load_only('type')
        ).first()
        if not terminal:
            self.query = self.query.filter(False)
            return
        terminal_query = m.db.session.query(m.SellableProductTerminal.sellable_product_id).filter(
            or_(
                m.SellableProductTerminal.terminal_code == terminal.code,
                and_(
                    m.SellableProductTerminal.apply_seller_id == current_user.seller_id,
                    m.SellableProductTerminal.terminal_id == 0,
                    m.SellableProductTerminal.terminal_type == terminal.type
                )
            )
        )
        if on_off_status:
            status = m.Misc.query.filter(
                m.Misc.type == 'on_off_status',
                m.Misc.id == on_off_status
            ).first()
            terminal_query = terminal_query.filter(
                m.SellableProductTerminal.on_off_status == getattr(status, 'code', None)
            )

        self.query = self.query.filter(
            m.SellableProduct.id.in_(terminal_query.subquery())
        )

    def _apply_terminal_group_filter(self, terminal_group_code):
        sellable_ids = m.db.session.query(m.SellableProductTerminalGroup.sellable_product_id).filter(
            m.SellableProductTerminalGroup.terminal_group_code == terminal_group_code
        )

        self.query = self.query.filter(
            m.SellableProduct.id.in_(sellable_ids.subquery())
        )

    def _apply_bundle_filter(self, is_bundle):
        self.query = self.query.filter(
            m.SellableProduct.is_bundle == is_bundle
        )

    def _apply_group_model_filter(self, models):
        _models = []
        for model in models:
            if model.strip():
                _models.append(model)
        self.query = self.query.filter(
            m.SellableProduct.model.in_(_models)
        )

    def _apply_restrict_convert_qty_filter(self, ):
        """
        restrict_convert_qty
        if restrict_convert_qty = 1/True:
            do not return skus which have uom_ratio = 1 and need_convert_qty = 1

        -> condition: where uom_ratio != 1 or need_convert_qty != 1
        """
        self.query = self.query.filter(
            or_(
                m.SellableProduct.uom_ratio != 1,
                m.SellableProduct.need_convert_qty != 1
            )
        )


class SubSkuListQuery(QueryBase):
    model = m.SellableProductSubSku

    def apply_filters(self, **filters):
        skus = filters.get('skus', [])
        self._filler_sub_skus(skus)

    def _filler_sub_skus(self, sub_skus):
        self.query = self.query.filter(
            m.SellableProductSubSku.sub_sku.in_(sub_skus)
        )


def _get_variant_attribute_sets(sellables):
    """
    :return: dict,{
        "attribute set id": "number of variant attribute"
    }
    """
    attribute_set_ids = set(map(lambda x: x.attribute_set_id, sellables))
    q = m.db.session.query(
        m.AttributeGroup.attribute_set_id,
        func.count(m.AttributeGroupAttribute.id)
    ).join(
        m.AttributeGroup,
        m.AttributeGroup.id == m.AttributeGroupAttribute.attribute_group_id
    ).filter(
        m.AttributeGroup.attribute_set_id.in_(attribute_set_ids),
        m.AttributeGroupAttribute.is_variation.is_(True)
    ).group_by(m.AttributeGroup.attribute_set_id).all()
    return dict(q)


def get_sellable_products(params, restrict_seller=True):
    """

    :param restrict_seller:
    :param params:
    :return:
    """
    query = SellableProductListQuery()
    query.apply_filters(restrict_seller, **params)
    total = len(query)
    page = params.get('page')
    page_size = params.get('page_size')
    query.query = query.query.options(
        load_only('id', 'name', 'product_id', 'sku', 'variant_id', 'seller_id', 'is_bundle', 'barcode',
                  'attribute_set_id', 'provider_id', 'seller_sku', 'uom_code', 'uom_name', 'uom_ratio'),
        joinedload('category').load_only('code', 'name'),
        joinedload('brand').load_only('id', 'name'),
        joinedload('editing_status').load_only('code', 'name', 'config'),
        joinedload('selling_status').load_only('code', 'name', 'config'),
        joinedload('product').load_only('name'),
    )
    query.pagination(page, page_size)
    items = query.all()
    variant_attribute_sets = _get_variant_attribute_sets(items)

    for item in items:
        flag = item.is_bundle or not item.attribute_set_id in variant_attribute_sets
        setattr(item, 'is_allow_create_variant', flag)

    return {
        'current_page': page,
        'page_size': page_size,
        'totalRecords': total,
        'skus': items
    }


def get_skus_uom_info(params, restrict_seller=False):
    """

    :param restrict_seller:
    :param params:
    :return:
    """
    query = SellableProductListQuery()
    query.apply_filters(**params)
    total = len(query)
    page = params.get('page', 1)
    page_size = params.get('page_size', 10)
    query.query = query.query.options(
        load_only('id', 'sku', 'seller_sku', 'seller_id', 'uom_code', 'uom_ratio', 'name', 'uom_name',
                  'need_convert_qty'),
    )
    query.pagination(page, page_size)
    items = query.all()

    return {
        'current_page': page,
        'page_size': page_size,
        'total_records': total,
        'skus': items
    }


def update_sellable_editing_status(
        status, ids=[], skus=[], sellables=None, comment=None, auto_commit=True, updated_by=None):
    if not sellables:
        if ids:
            sellables = m.SellableProduct.query.filter(
                m.SellableProduct.id.in_(ids)
            )
        else:
            sellables = m.SellableProduct.query.filter(
                m.SellableProduct.sku.in_(skus)
            )
    for sellable in sellables:
        variant = sellable.product_variant
        product = variant.product

        if status == 'pending_approval':
            if product.editing_status_code != 'active':
                product.editing_status_code = 'pending_approval'
            if variant.editing_status_code != 'active':
                variant.editing_status_code = 'pending_approval'
            sellable.editing_status_code = 'pending_approval'
        elif status == 'active':
            product.editing_status_code = 'active'
            variant.editing_status_code = 'active'
            sellable.editing_status_code = 'active'
        elif status == 'reject':
            if product.editing_status_code != 'active':
                product.editing_status_code = 'reject'
            if variant.editing_status_code != 'active':
                variant.editing_status_code = 'reject'
            sellable.editing_status_code = 'reject'
        elif status == 'processing':
            sellable.editing_status_code = 'processing'
        elif status == 'inactive':
            if product.editing_status_code != 'active':
                product.editing_status_code = 'inactive'
            if variant.editing_status_code != 'active':
                variant.editing_status_code = 'inactive'
            sellable.editing_status_code = 'inactive'
        elif status == 'suspend':
            sellable.editing_status_code = 'suspend'

        if comment is not None:
            sellable.comment = comment
        sellable.updated_by = updated_by or current_user.email

    if auto_commit:
        m.db.session.commit()
        for item in sellables:
            signals.sellable_common_update_signal.send(item)
            signals.sellable_update_signal.send(item)
    return sellables


def update_sellable_product_tag(sellable_product_id, sku, tags, overwrite, auto_commit=True):
    new_tags = list(set(tags.split(",")))

    product_tag = m.SellableProductTag.query.filter(
        m.SellableProductTag.sellable_product_id == sellable_product_id
    ).first()

    if not product_tag:
        product_tag = m.SellableProductTag(
            sellable_product_id=sellable_product_id,
            sku=sku,
            tags=tags,
            created_by=current_user.email,
            updated_by=current_user.email
        )
        new_tags_set = new_tags
    else:
        if overwrite.upper() == 'Y':
            new_tags_set = new_tags
        else:
            old_tags = product_tag.tags.split(",")
            new_tags_set = list(set(old_tags + new_tags))
    new_tags_set = list(filter(None, [t.strip() if t.strip() else None for t in new_tags_set]))
    tags = ",".join(new_tags_set)

    if len(new_tags_set) > 20 or len(list(filter(lambda x: len(x) > 60, new_tags_set))) > 0:
        raise ValueError("Không quá 20 tags 1 sản phẩm, và 1 tag không quá 60 ký tự")

    product_tag.tags = tags
    product_tag.updated_by = current_user.email

    m.db.session.add(product_tag)
    m.db.session.flush()

    if auto_commit:
        m.db.session.commit()

    return product_tag


def update_terminal_groups_for_sellable(sellable_product_id, terminal_groups_new, terminal_groups_delete):
    # get current teminal group codes for sellable product
    tmp_terminal_group_codes = get_terminal_groups_of_sellable_product(sellable_product_id)  # get_terminal_groups()

    # Get terminal groups passed by import file and put in set to remove duplicated
    new_terminal_groups = []
    delete_teminal_groups = set()
    if terminal_groups_new != None:
        new_terminal_groups = list(set(tg.strip().split('=>')[0] for tg in terminal_groups_new.split(",") if tg))
    if terminal_groups_delete != None:
        delete_teminal_groups = set(tg.strip().split('=>')[0] for tg in terminal_groups_delete.split(",") if tg)
    tmp_terminal_group_codes = set(new_terminal_groups + tmp_terminal_group_codes)
    tmp_terminal_group_codes -= delete_teminal_groups

    terminal_groups = [code for code in tmp_terminal_group_codes if code]

    payload = {
        'sellable_products': [sellable_product_id],
        'terminal_groups': terminal_groups
    }

    sellable_validator.UpsertSellableProductTerminalGroup.validate(payload)

    # upsert terminal groups for sellable
    upsert_sellable_product_terminal_group([sellable_product_id], terminal_groups)


def update_common(sku_id=None, sku=None, data=None, autocommit=True, overwrite_barcode=True,
                  sellable_common_update_signal=True, sellable_update_signal=True):
    if sku_id:
        sellable = m.SellableProduct.query.filter(
            m.SellableProduct.id == sku_id).first()
    else:
        sellable = m.SellableProduct.query.filter(
            m.SellableProduct.sku == sku).first()

    updated_by = data.get('created_by') or current_user.email
    for key, value in data.items():
        if key == 'manage_serial' and not data.get('tracking_type'):
            sellable.tracking_type = value
        elif key == 'auto_generate_serial':
            manage_serial = sellable.manage_serial
            if manage_serial is not None:
                if manage_serial:
                    sellable.auto_generate_serial = value
                else:
                    sellable.auto_generate_serial = False

        elif key == 'shipping_types':
            shipping_types = data.get('shipping_types')
            if shipping_types is not None:
                SellableProductShippingTypeService.delete(sellable.id, auto_commit=False)
                for shipping_type_id in shipping_types:
                    SellableProductShippingTypeService.create(sellable.id, shipping_type_id, auto_commit=False)

        elif key == 'type':
            sellable.product_type = value
        elif key == 'description':
            sellable.terminal_seo.short_description = value
        elif key == 'detailed_description':
            sellable.terminal_seo.description = value
        elif key == 'editing_status_code':
            update_sellable_editing_status(status=value, sellables=[sellable], updated_by=updated_by)
        elif key == 'barcodes':
            if overwrite_barcode is True:
                _delete_sellable_product_barcodes(sellable.id)
                sellable.barcode = None
            if value:
                _insert_sellable_product_barcodes(sellable.id, value, data.get('created_by'),
                                                  update=not overwrite_barcode)
                sellable.barcode = value[-1].get('barcode')
        elif key == 'model':
            _update_model_sellable_product(sellable, data.get('model'))
        elif key == 'category_id':
            _update_category_sellable_product(sellable, data.get('category_id'), data.get('created_by'))
        else:
            setattr(sellable, key, value)

    sellable.updated_by = updated_by

    m.db.session.flush()
    has_shipping_type = m.db.session.query(exists().where(
        m.SellableProductShippingType.sellable_product_id == sellable.id)).scalar()
    if not has_shipping_type:
        default_shipping_type = get_default_shipping_type()
        if default_shipping_type:
            SellableProductShippingTypeService.create(
                sellable_product_id=sellable.id,
                shipping_type_id=default_shipping_type.id,
                auto_commit=False
            )
    if autocommit:
        m.db.session.commit()

    sellable_query = m.SellableProduct.query
    if sku_id:
        sellable_query = sellable_query.filter(m.SellableProduct.id == sku_id)
    else:
        sellable_query = sellable_query.filter(m.SellableProduct.sku == sku)

    if sellable_common_update_signal or sellable_update_signal:
        sellable = sellable_query.options(
            joinedload('unit'),
            joinedload('unit_po'),
            joinedload('category'),
            joinedload('brand'),
        ).first()

    if sellable_common_update_signal:
        signals.sellable_common_update_signal.send(sellable)
    if sellable_update_signal:
        signals.sellable_update_signal.send(sellable)
    return sellable_query.first()


def move_skus_to_single_product(skus, product_id):
    existing_sku_obj = m.SellableProduct.query.filter(
        m.SellableProduct.product_id == product_id,
    ).first()

    target_variant_attrs = m.VariantAttribute.query.join(
        m.AttributeGroupAttribute,
        m.AttributeGroupAttribute.attribute_id == m.VariantAttribute.attribute_id
    ).join(
        m.AttributeGroup,
        m.AttributeGroup.id == m.AttributeGroupAttribute.attribute_group_id
    ).join(
        m.AttributeSet,
        m.AttributeSet.id == m.AttributeGroup.attribute_set_id
    ).filter(
        m.VariantAttribute.variant_id == existing_sku_obj.variant_id,
        m.AttributeGroupAttribute.is_variation == True,
        m.AttributeSet.id == existing_sku_obj.attribute_set_id
    ).order_by(
        m.VariantAttribute.id,
        m.VariantAttribute.attribute_id
    ).all()

    variation_attr_ids = list(map(lambda x: x.attribute_id, target_variant_attrs))

    target_product = m.Product.query.get(product_id)
    sellable_products = m.SellableProduct.query.filter(
        m.SellableProduct.sku.in_(skus)
    ).all()

    product_ids = list(map(lambda x: x.product_id, sellable_products))
    product_ids = list(set(product_ids))
    product_ids.append(product_id)

    moving_variant_ids = []
    skus_need_update_on_product_details = []
    for sellable_product in sellable_products:
        moving_variant_ids.append(sellable_product.variant_id)
        sellable_product.product_id = product_id
        sellable_product.attribute_set_id = target_product.attribute_set_id
        sellable_product.category_id = target_product.category_id

    current_varriants = m.ProductVariant.query.filter(
        m.ProductVariant.id.in_(moving_variant_ids)
    ).all()

    for current_variant in current_varriants:
        current_variant.product_id = product_id

    target_normal_attrs = m.VariantAttribute.query.options(
        load_only(m.VariantAttribute.attribute_id, m.VariantAttribute.value, m.VariantAttribute.unit_id)
    ).filter(
        m.VariantAttribute.variant_id == existing_sku_obj.variant_id,
        m.VariantAttribute.attribute_id.notin_(variation_attr_ids)
    ).all()

    moving_normal_attribute = m.VariantAttribute.query.options(
        load_only(m.VariantAttribute.attribute_id, m.VariantAttribute.value)
    ).filter(
        m.VariantAttribute.variant_id.in_(moving_variant_ids),
        m.VariantAttribute.attribute_id.notin_(variation_attr_ids)
    ).all()

    target_normal_attrs_dict = {x.attribute_id: x for x in target_normal_attrs}

    moving_normal_attribute_ids = list(map(
        lambda x: x.attribute_id, moving_normal_attribute
    ))

    remove_variant_attr_ids = list(filter(
        lambda x: x not in target_normal_attrs_dict.keys(), moving_normal_attribute_ids
    ))

    insert_variant_attr_ids = list(filter(
        lambda x: x not in moving_normal_attribute_ids, target_normal_attrs_dict.keys()
    ))
    remove_variant_attr_ids = list(set(remove_variant_attr_ids))
    insert_variant_attr_ids = list(set(insert_variant_attr_ids))
    if remove_variant_attr_ids:
        m.VariantAttribute.query.filter(
            m.VariantAttribute.variant_id.in_(moving_variant_ids),
            m.VariantAttribute.attribute_id.in_(remove_variant_attr_ids)
        ).delete(synchronize_session='fetch')

    if insert_variant_attr_ids:
        for insert_variant_attr_id in insert_variant_attr_ids:
            for moving_variant_id in moving_variant_ids:
                new_moving_variant_attr = m.VariantAttribute(
                    variant_id=moving_variant_id,
                    attribute_id=insert_variant_attr_id,
                    value=target_normal_attrs_dict.get(insert_variant_attr_id).value,
                    unit_id=target_normal_attrs_dict.get(insert_variant_attr_id).unit_id
                )
                m.db.session.add(new_moving_variant_attr)

    m.db.session.commit()

    if existing_sku_obj:  # publish event AddVariantSkuMsg to Clearance svc
        from catalog.extensions.ram.publisher import add_variant_sku_ram_publisher
        from catalog.extensions.ram.publisher import AddVariantSkuMsg
        for sku in skus:
            add_variant_sku_ram_publisher.publish(AddVariantSkuMsg(
                variant_sku=sku,
                sibling_sku=existing_sku_obj.sku,
            ))

    filterd_product_ids = []
    sellable_product_of_product_ids = SellableProduct.query.filter(
        SellableProduct.product_id.in_(product_ids)
    ).all()

    for sellable_product_of_product_id in sellable_product_of_product_ids:
        if not sellable_product_of_product_id.product_id in filterd_product_ids:
            signals.sellable_update_signal.send(sellable_product_of_product_id)
            filterd_product_ids.append(sellable_product_of_product_id.product_id)

    return sellable_products


def get_sellables_for_export(**params):
    attribute_set_id = params.get('params', {}).get('attribute_set')
    if params.get('include_attribute'):
        if not attribute_set_id:
            raise BadRequestException('missing attribute_set param')
        elif ',' in str(attribute_set_id):
            raise BadRequestException('only 1 attribute_set_id is allow at attribute_set param')
        del params['include_attribute']
    query = base_export_query(params.get('params'))
    n_record = len(query)
    if n_record > MAX_RECORD:
        raise exc.BadRequestException(f'Có {n_record} kết quả, vượt quá {MAX_RECORD} bản ghi')
    params.update({'email': current_user.email})
    signals.export_product_signal.send(params)


def update_item_for_bundle_sellable(bundle_id, items):
    bundle = m.SellableProduct.query.get(bundle_id)
    bundle.children.clear()
    m.db.session.flush()
    for index, item_info in enumerate(items):
        item = m.SellableProductBundle()
        item.bundle_id = bundle.id
        item.sellable_product_id = item_info['id']
        item.quantity = item_info['quantity']
        item.priority = index
        m.db.session.add(item)
    m.db.session.commit()

    for item in items:
        bundle_item = m.SellableProduct.query.get(item['id'])
        signals.sellable_update_signal.send(bundle_item)
    signals.sellable_update_signal.send(bundle)


def _safe_get_value(obj, name):
    if hasattr(obj, name):
        return getattr(obj, name)
    return None


def _transform_sku_to_dict(item, yes_no_map, attributes=[]):
    result = {
        'category code': _safe_get_value(item.line_category, 'code'),
        'category name': _safe_get_value(item.line_category, 'name'),
        'master category name': _safe_get_value(item.master_category, 'name'),
        'product name': item.name,
        'sku': item.sku,
        'seller sku': item.seller_sku,
        'brand': _safe_get_value(item.brand, 'name'),
        'model': item.model,
        'warranty months': item.warranty_months,
        'warranty note': item.warranty_note,
        'uom': item.uom_name,
        'uom_ratio': item.uom_ratio,
        'vendor tax': _safe_get_value(item.tax_in, 'label'),
        'short description': _safe_get_value(item.terminal_seo, 'short_description'),
        'description': _safe_get_value(item.terminal_seo, 'description'),
        'part number': item.part_number,
        'barcode': item.barcode,
        'allow selling without stock?': yes_no_map.get(item.allow_selling_without_stock, 'Yes'),
        'is tracking serial?': yes_no_map.get(item.tracking_type, 'Yes'),
        'expiry tracking': yes_no_map.get(item.expiry_tracking, 'Yes'),
        'expiration type': item.get_expiration_type(),
        'days before Exp lock': item.days_before_exp_lock,
    }
    variant_attributes = item.get_attributes_by_codes([attribute.code for attribute in attributes])
    for variant_attribute in variant_attributes:
        result[variant_attribute.attribute.code] = variant_attribute.value
    return result


def sellables_exporter(query):
    attributes = m.Attribute.query.filter(
        m.Attribute.code.in_(DIMENSION_ATTRIBUTES_CODES + PACK_CODE_ATTRIBUTES)
    ).all()

    items = query.all()
    # pylint: disable=logging-too-many-args
    template_filename = os.path.join(
        ROOT_DIR,
        'storage',
        'template',
        'template_export_product_v2.0.xlsx'
    )
    START_ROW = 3
    yes_no_map = {
        0: 'No',
        None: '',
    }

    wb = openpyxl.load_workbook(template_filename)
    ws = wb['SanPham']
    column_map = {}
    for index, (col, *_) in enumerate(ws.iter_cols(1, 30, 2, 2)):
        column_map[col.value] = index + 1
    for row_idx, item in enumerate(items, START_ROW):
        data = {
            'no': row_idx - 2,
            **_transform_sku_to_dict(item, yes_no_map, attributes)
        }
        # fill data to row
        for key, value in data.items():
            col_idx = column_map.get(key)
            if col_idx:
                try:
                    ws.cell(row_idx, col_idx, value)
                except Exception as ex:
                    _logger.exception(
                        'Type of a cell in the excel file:',
                        value,
                        ex
                    )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def seo_info_exporter(items):
    # pylint: disable=logging-too-many-args
    template_filename = os.path.join(
        ROOT_DIR,
        'storage',
        'template',
        'template_export_seo_info.xlsx'
    )
    START_ROW = 3

    wb = openpyxl.load_workbook(template_filename)
    ws = wb['SanPham']
    column_map = {}
    for index, (col, *_) in enumerate(ws.iter_cols(1, 30, 2, 2)):
        column_map[col.value] = index + 1
    for row_idx, item in enumerate(items, START_ROW):
        data = {
            'seller_sku': item.seller_sku,
            'uom_code': item.uom_code,
            'uom_ratio': item.uom_ratio,
            'display_name': _safe_get_value(item.terminal_seo, 'display_name'),
            'meta_title': _safe_get_value(item.terminal_seo, 'meta_title'),
            'meta_description': _safe_get_value(item.terminal_seo, 'meta_description'),
            'meta_keyword': _safe_get_value(item.terminal_seo, 'meta_keyword'),
            'url_key': _safe_get_value(item.terminal_seo, 'url_key'),
        }
        # fill data to row
        for key, value in data.items():
            col_idx = column_map.get(key)
            if col_idx:
                try:
                    ws.cell(row_idx, col_idx, value)
                except Exception as ex:
                    _logger.exception(
                        'Type of a cell in the excel file:',
                        value,
                        ex
                    )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _transform_detail_sku_to_dict(item, yes_no_map, uom_map):
    return {
        'sku': item.sku,
        'seller sku': item.seller_sku,
        'category': _safe_get_value(item.line_category, 'code') + '=>' + _safe_get_value(
            item.line_category, 'name') if item.line_category else '',
        'master category': _safe_get_value(item.master_category, 'code'),
        'product name': item.name,
        'brand': _safe_get_value(item.brand, 'name'),
        'model': item.model,
        'warranty months': item.warranty_months,
        'warranty note': item.warranty_note,
        'uom': item.uom_name,
        'uom_ratio': item.uom_ratio,
        'vendor tax': _safe_get_value(item.tax_in, 'label'),
        'short description': _safe_get_value(item.terminal_seo, 'short_description'),
        'description': _safe_get_value(item.terminal_seo, 'description'),
        'part number': item.part_number,
        'barcode': item.barcode,
        'allow selling without stock?': yes_no_map.get(item.allow_selling_without_stock, 'Yes'),
        'is tracking serial?': yes_no_map.get(item.tracking_type, 'Yes'),
        'expiry tracking': yes_no_map.get(item.expiry_tracking, 'Yes'),
        'expiration type': item.get_expiration_type(),
        'days before Exp lock': item.days_before_exp_lock,
    }


def detail_sellables_exporter(items, attribute_set_id):
    template_filename = os.path.join(ROOT_DIR,
                                     'storage',
                                     'template',
                                     'template_export_detail_product_v1.0.xlsx'
                                     )

    if not attribute_set_id:
        raise BadRequestException('missing attribute_set_id')
    ATTRIBUTE_COLUMN_INDEX = 22
    items = items.all()
    START_ROW = 3
    yes_no_map = {
        0: 'No',
        None: '',
    }

    wb = openpyxl.load_workbook(template_filename)
    ws = wb['SanPham']
    column_map = {}
    for index, (col, *_) in enumerate(ws.iter_cols(1, 30, 2, 2)):
        column_map[col.value] = index + 1

    # attribute header
    attribute_query = text("select a.id, a.code, a.name from attributes a "
                           "join attribute_group_attribute aga on a.id = aga.attribute_id "
                           "join attribute_groups ag on aga.attribute_group_id = ag.id "
                           "where ag.attribute_set_id = :attribute_set_id and ag.system_group != 1 "
                           "order by ag.priority asc, aga.priority asc")
    rs = db.engine.execute(attribute_query, {
        'attribute_set_id': attribute_set_id
    })
    attribute_headers = [{
        "id": row[0],
        "code": row[1],
        "name": row[2]
    } for row in rs]

    name_header_row_idx = 0
    code_header_row_idx = 1
    for index, attribute in enumerate(attribute_headers):
        ws.cell(row=name_header_row_idx + 1, column=index + ATTRIBUTE_COLUMN_INDEX + 1).value = attribute["name"]
        ws.cell(row=code_header_row_idx + 1, column=index + ATTRIBUTE_COLUMN_INDEX + 1).value = attribute["code"]

    variant_ids = LambdaList(items).map(lambda x: x.variant_id).list()

    product_attributes = m.VariantAttribute.query.filter(
        m.VariantAttribute.variant_id.in_(variant_ids)).all()
    dic_variant_id_product_attributes = {}
    for product_attribute in product_attributes:
        if product_attribute.variant_id in dic_variant_id_product_attributes:
            dic_variant_id_product_attributes[product_attribute.variant_id].append(product_attribute)
        else:
            dic_variant_id_product_attributes[product_attribute.variant_id] = [product_attribute]

    attribute_ids = LambdaList(product_attributes).map(lambda x: x.attribute_id).list()
    attributes = m.Attribute.query.filter(
        m.Attribute.id.in_(attribute_ids)).all()
    dic_attribute = {x.id: x for x in attributes}

    attribute_option_ids = []
    dic_pa_id_ao_ids = {}
    for product_attribute in product_attributes:
        attribute = dic_attribute[product_attribute.attribute_id]
        if attribute.value_type == AttributeValueType.SELECTION.value:
            ao_id = int(product_attribute.value)
            attribute_option_ids.append(ao_id)
            dic_pa_id_ao_ids[product_attribute.id] = [ao_id]
        elif attribute.value_type == AttributeValueType.MULTI_SELECT.value:
            ao_ids = LambdaList(str(product_attribute.value).split(',')).map(lambda x: int(x)).list()
            attribute_option_ids += ao_ids
            dic_pa_id_ao_ids[product_attribute.id] = ao_ids

    attribute_options = m.AttributeOption.query.filter(
        m.AttributeOption.id.in_(attribute_option_ids)).all()
    dic_attribute_option = {x.id: x for x in attribute_options}

    dic_sellable_id_attribute_value = {}
    uom_map = {}
    for item in items:
        s_product_attributes = dic_variant_id_product_attributes.get(item.variant_id, [])
        dic_s_attribute_value = {}
        for s_product_attribute in s_product_attributes:
            s_attribute = dic_attribute.get(s_product_attribute.attribute_id)
            if not s_attribute:
                pass
            elif s_attribute.value_type == AttributeValueType.SELECTION.value \
                    or s_attribute.value_type == AttributeValueType.MULTI_SELECT.value:
                s_ao_ids = dic_pa_id_ao_ids.get(s_product_attribute.id, [])
                s_value = LambdaList(s_ao_ids).map(lambda x: dic_attribute_option.get(x)) \
                    .filter(lambda x: x).map(lambda x: x.value).string_join(',')
                dic_s_attribute_value[s_attribute.id] = s_value
            else:
                dic_s_attribute_value[s_attribute.id] = s_product_attribute.value
            dic_sellable_id_attribute_value[item.id] = dic_s_attribute_value
            if s_attribute.code == UOM_CODE_ATTRIBUTE:
                uom_attribute = dic_attribute_option.get(int(s_product_attribute.value))
                uom_map[item.id] = uom_attribute.value if uom_attribute else ''

    for row_idx, item in enumerate(items, START_ROW):
        data = {
            'no': row_idx - 2,
            **_transform_detail_sku_to_dict(item, yes_no_map, uom_map)
        }
        # fill data to row
        for key, value in data.items():
            col_idx = column_map.get(key)
            if col_idx:
                ws.cell(row_idx, col_idx, value)
            dic_s_attribute_value = dic_sellable_id_attribute_value.get(item.id, {})
            for idx_attr, attribute_header in enumerate(attribute_headers):
                ws.cell(row_idx, idx_attr + ATTRIBUTE_COLUMN_INDEX + 1,
                        dic_s_attribute_value.get(attribute_header['id'], ''))

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def base_export_query(params):
    wrapper_query = SellableProductListQuery()
    wrapper_query.apply_filters(**params, restrict_seller=False)
    return wrapper_query


def sellables_export_query_builder(params, export_type=None):
    wrapper_query = base_export_query(params)
    query = wrapper_query.query
    if export_type == ExportSellable.EXPORT_SEO_INFO:
        query = query.options(
            joinedload('terminal_seo').load_only(
                'display_name',
                'meta_title',
                'meta_description',
                'meta_keyword',
                'url_key'
            ),
            noload('selling_status'),
            noload('editing_status'),
            noload('product'),
            noload('product_variant'),
            noload('seller'),
            load_only(
                'seller_sku',
                'uom_code',
                'uom_ratio'
            )
        )
    else:
        query = query.options(
            joinedload('master_category').load_only('name'),
            joinedload('tax_in').load_only('label'),
            joinedload('brand').load_only('name'),
            joinedload('terminal_seo').load_only('description', 'short_description'),

            noload('selling_status'),
            noload('editing_status'),
            noload('product'),
            noload('product_variant'),
            noload('seller'),

            load_only(
                'sku',
                'seller_sku',
                'name',
                'model',
                'warranty_note',
                'warranty_months',
                'part_number',
                'barcode',
                'allow_selling_without_stock',
                'tracking_type',
                'auto_generate_serial',
                'expiry_tracking',
                'expiration_type',
                'days_before_exp_lock',
                'uom_code',
                'uom_ratio'
            ),
        )
    return query


def get_sellable_products_in_bundle(bundle_id):
    items = []
    bundle = m.SellableProductBundle.query.filter(
        m.SellableProductBundle.bundle_id == bundle_id
    ).order_by(m.SellableProductBundle.priority).all()

    for item in bundle:
        if item.sellable_product is None:
            continue

        editing_status = m.EditingStatus.query.filter(
            m.EditingStatus.code == item.sellable_product.editing_status_code
        ).with_entities(m.EditingStatus.code, m.EditingStatus.name, m.EditingStatus.config).first()

        selling_status = m.SellingStatus.query.filter(
            m.SellingStatus.code == item.sellable_product.selling_status_code
        ).with_entities(m.SellingStatus.code, m.SellingStatus.name, m.SellingStatus.config).first()

        sellalbe_product = {
            'id': item.sellable_product.id,
            'sku': item.sellable_product.sku,
            'name': item.sellable_product.name,
            'editing_status': editing_status,
            'quantity': item.quantity,
            'priority': item.priority,
            'selling_status': selling_status,
            'allow_selling_without_stock': item.sellable_product.allow_selling_without_stock
        }
        items.append(sellalbe_product)

    return items


def update_status_from_srm(sellable_id, srm_code):
    sellable = m.SellableProduct.query.get(sellable_id)
    if srm_code:
        srm_status = m.SRMStatus.query.filter(
            m.SRMStatus.code == srm_code
        ).first()
        sellable.selling_status_code = srm_status.selling_status
    else:
        sellable.selling_status_code = None
    m.db.session.commit()
    return sellable


def get_seo_info_of_sellable_product_on_terminal(sellable_id=None, sku=None, terminal_codes=None):
    if not sellable_id:
        sellable_id = m.SellableProduct.query.filter_by(
            sku=sku
        ).first().id

    product = m.Product.query.filter(
        m.Product.id == m.SellableProduct.product_id,
        m.SellableProduct.id == sellable_id
    ).options(load_only('display_name', 'meta_title', 'meta_description', 'meta_keyword', 'url_key')).first()

    return product


def upsert_info_of_sellable_product_on_terminals(seo_info, terminal_codes=None, sellable_id=None, sku=None):
    if not sellable_id:
        sellable_product = m.SellableProduct.query.filter(m.SellableProduct.sku == sku).first()
        sellable_id = sellable_product.id
    else:
        sellable_product = m.SellableProduct.query.filter(m.SellableProduct.id == sellable_id).first()

    product = m.Product.query.filter(m.Product.id == sellable_product.product_id).first()

    product.display_name = seo_info.get('display_name', product.display_name)
    product.url_key = seo_info.get('url_key', product.url_key)
    product.meta_title = seo_info.get('meta_title', product.meta_title)
    product.meta_description = seo_info.get('meta_description', product.meta_description)
    product.meta_keyword = seo_info.get('meta_keyword', product.meta_keyword)

    seo = m.SellableProductSeoInfoTerminal.query.filter(
        m.SellableProductSeoInfoTerminal.sellable_product_id == sellable_id,
        m.SellableProductSeoInfoTerminal.terminal_id == 0
    ).first()

    if seo:
        seo.description = seo_info.get('description', seo.description)
        seo.short_description = seo_info.get('short_description', seo.short_description)

        seo.updated_by = current_user.email
    else:
        seo = m.SellableProductSeoInfoTerminal()

        seo.description = seo_info.get('description')
        seo.short_description = seo_info.get('short_description')
        seo.terminal_id = 0

        seo.sellable_product_id = sellable_id
        seo.created_by = current_user.email
        seo.updated_by = current_user.email

        db.session.add(seo)

    db.session.commit()
    signals.sellable_update_seo_info_signal.send(sellable_product)


def get_terminal_groups_of_sellable_product(sellable_id):
    sellable_terminal_groups = m.SellableProductTerminalGroup.query.join(
        m.SellableProduct,
        m.SellableProduct.id == m.SellableProductTerminalGroup.sellable_product_id
    ).filter(
        m.SellableProduct.seller_id == current_user.seller_id,
        m.SellableProduct.id == sellable_id
    ).all()

    return [sellable_terminal_group.terminal_group_code for sellable_terminal_group in sellable_terminal_groups]


def upsert_sellable_product_terminal_group(sellable_products, terminal_groups):
    """

    delete all old (seller_id, sellable_product_id, terminal_group_code) records
    create and insert new SellableProductTerminalGroup records
    :param sellable_products list<int>
    :param terminal_groups list<string>
    """

    m.SellableProductTerminalGroup.query.filter(
        m.SellableProductTerminalGroup.sellable_product_id.in_(sellable_products)
    ).delete(False)
    m.db.session.commit()

    data = []
    for sellable_id in sellable_products:
        for terminal_group_code in terminal_groups:
            data.append(
                {"sellable_product_id": sellable_id,
                 "terminal_group_code": terminal_group_code,
                 "created_by": current_user.email,
                 "updated_by": current_user.email})

    if data:
        insert_query = m.db.insert(m.SellableProductTerminalGroup).values(data)
        m.db.session.execute(insert_query)
        m.db.session.commit()

    active_sellable_products = m.SellableProduct.query.filter(
        m.SellableProduct.id.in_(sellable_products)
    ).all()

    for sellable in active_sellable_products:
        signals.sellable_update_signal.send(sellable)


def get_sellable_product(sp_id):
    return m.SellableProduct.query.filter(m.SellableProduct.id == sp_id).first()


def get_skus_by_filter(seller_id, seller_sku, uom_name=None, uom_ratio=None, only_one=False):
    query = m.SellableProduct.query.filter(and_(m.SellableProduct.seller_sku == seller_sku,
                                                m.SellableProduct.seller_id == seller_id))
    if uom_name:
        uom_code = AttributeService.get_uom_code_by_name(seller_id, uom_name)
        if not uom_code:
            raise BadRequestException('Không tìm thấy sản phẩm (kiểm tra lại mã seller sku, '
                                      'đơn vị tính và tỷ lệ quy đổi)')
        query = query.filter(m.SellableProduct.uom_code == uom_code)
    if uom_ratio is not None and uom_ratio != '':
        query = query.filter(m.SellableProduct.uom_ratio == uom_ratio)
    skus = query.all()
    if only_one:
        if not skus:
            raise BadRequestException('Không tìm thấy sản phẩm (kiểm tra lại mã seller sku, '
                                      'đơn vị tính và tỷ lệ quy đổi)')
        if len(skus) >= 2:
            raise BadRequestException('Tìm thấy nhiều hơn 1 sản phẩm có cùng mã seller sku. '
                                      'Vui lòng nhập thêm đơn vị tính và tỷ lệ quy đổi')
        return skus[0]
    return skus
