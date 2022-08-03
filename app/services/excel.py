import os
import openpyxl as xl
import config
import funcy

from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

from catalog import models as m, services

from catalog.extensions import exceptions as excs

PRODUCT_EXPORT_FIELDS = [
    'sku', 'name', 'seller_sku', 'category_name', 'sale_category_name',
    'brand_name', 'color_name', 'model', 'part_number', 'barcode',
    'unit_name', 'uom_po_name', 'type_name', 'objective_name', 'warranty_months',
    'warranty_description', 'weight', 'length', 'width',
    'height', 'tax_in', 'tax_out', 'sale_channels_name', 'product_display_name',
    'short_description', 'meta_title', 'meta_keyword', 'meta_description'

]

MAX_PRODUCTS_EXPORT = 1000


def export_product_list(**kwargs):
    if 'attributeSetId' in kwargs and kwargs['attributeSetId'] is not None:
        if not kwargs['exportCommon'] and len(kwargs['attributeSetId'].split(',')) != 1:
            raise excs.BadRequestException('Chỉ đuợc chọn 1 bộ thuộc tính')
    product_list_query = ProductExportListQuery()
    product_list_query.apply_filters(**kwargs)
    n_products = len(product_list_query)
    if not (0 < n_products <= MAX_PRODUCTS_EXPORT):
        raise excs.UnprocessableEntityException(
            f'Chỉ export khi danh sách sản phẩm có số lượng không quá {MAX_PRODUCTS_EXPORT}')

    return export_product(kwargs)


def add_product(product, sheet, row):
    """
    :type product_model: m.Product
    :param product:
    :param sheet:
    :return:
    """

    sheet.cell(row=row, column=1).value = row - 3
    sheet.cell(row=row, column=2).value = 'Bundle' if product.is_bundle else 'Simple'
    column = 2
    for field in PRODUCT_EXPORT_FIELDS:
        column += 1
        if hasattr(product, field):
            sheet.cell(row=row, column=column).value = getattr(product, field)
            continue
        sheet.cell(row=row, column=column).value = ""


def add_title_common(sheet, attributes):
    for i, attribute in enumerate(attributes):
        column = 3 + len(PRODUCT_EXPORT_FIELDS) + i
        sheet.cell(row=3, column=column).value = attribute.name
    return


def remove_header_common(sheet):
    column = 3 + len(PRODUCT_EXPORT_FIELDS)
    sheet.cell(row=2, column=column).value = ""
    return


def add_common(sheet, attributes, product, row):
    """
    :type product: ProductExport
    :param row:
    :param sheet:
    :param attributes:
    :param product:
    :return:
    """
    for i, attribute in enumerate(attributes):
        column = 3 + len(PRODUCT_EXPORT_FIELDS) + i
        sheet.cell(row=row, column=column).value = product.get_common_attribute(
            attribute
        )
    return


def export_product(kwargs):
    """
    @Todo
    :param kwargs:
    :return:
    """

    return
    product_export_list = ProductExportListQuery()
    product_export_list.apply_filters(**kwargs)
    products = product_export_list.query.limit(2000).all()
    file_path = os.path.join(
        config.ROOT_DIR,
        'storage',
        'template',
        'template_export_product.xlsx'
    )
    x = xl.load_workbook(file_path)
    sheet = x.get_sheet_by_name('SanPham')
    row = 4
    attributes = None

    set_id = kwargs.get('attributeSetId')
    export_common = kwargs.get('exportCommon')
    if export_common is False:
        attributes = get_attributes(set_id)
        add_title_common(sheet, attributes)
    else:
        remove_header_common(sheet)
    for product in products:
        add_product(product, sheet, row)

        if export_common is False:
            add_common(
                sheet=sheet,
                product=product,
                row=row,
                attributes=attributes
            )
        row += 1

    ws2 = x.get_sheet_by_name('DuLieuMau')
    fill_sample_data(ws2, set_id, export_common)

    return x, file_name()


def file_name():
    now = datetime.now()
    dt_string = now.strftime("%Y%m%d%H%M%S")
    return '{}_product.xlsx'.format(dt_string)


class ProductExportListQuery(object):
    def __init__(self):
        super().__init__()
        self.query = ProductExport.query


class ProductExport(m.Product):

    @property
    def category_name(self):
        if self.category:
            return self.category.name
        return None

    @property
    def color_name(self):
        if self.color:
            return self.color.name
        return None

    @property
    def sale_category_name(self):
        if self.sale_categories:
            category_names = []
            for sale_category in self.sale_categories:
                category_names.append(sale_category.sale_category.name)
            return ';'.join(category_names)
        return None

    @property
    def short_description(self):

        return None

    @property
    def meta_keyword(self):
        return None

    @property
    def meta_description(self):
        return None

    @property
    def meta_title(self):

        return None

    @property
    def sale_channels_name(self):
        if self.sale_channels:
            sale_channels_name = []
            for sale_channel in self.sale_channels:
                sale_channels_name.append(sale_channel.channel.name)
            return ';'.join(sale_channels_name)
        return None

    @property
    def objective_name(self):
        if self.objective_obj:
            return self.objective_obj.name
        return None

    @property
    def brand_name(self):
        if self.brand:
            return self.brand.name
        return None

    @property
    def tax_in(self):
        if self.property:
            return self.property.tax_in
        return None

    @property
    def tax_out(self):
        if self.property:
            return self.property.tax_out
        return None

    @property
    def product_display_name(self):
        return None

    def get_common_attribute(self, attribute):
        """
        :type attribute: m.Attribute
        :param attribute:
        :return:
        """
        product_attribute = m.ProductAttribute.query.filter(
            m.ProductAttribute.attribute_id == attribute.id,
            m.ProductAttribute.product_id == self.id
        ).first()  # type: m.ProductAttribute

        if product_attribute is None:
            return

        if attribute.value_type in ['text', 'number']:
            return product_attribute.value

        if attribute.value_type in ['selection', 'multiple_select'] and product_attribute.value:
            option_ids = str(product_attribute.value).split(",")
            options = m.AttributeOption.query.filter(
                m.AttributeOption.id.in_(option_ids)
            ).all()  # type: list[m.AttributeOption]
            option_values = funcy.lpluck_attr('value', options)
            return ';'.join(option_values)


def fill_sample_data(ws, attribute_set_id, export_common=True):
    """
    @Todo
    :param ws:
    :param attribute_set_id:
    :param export_common:
    :return:
    """
    return
    start_col = 15
    bg_color = 'CCC0DA'
    attributes_data, extra_data = services.import_data.get_data_for_excel_sheet(
        attribute_set_id
    )

    # Write default data to DuLieuMau sheet
    default_cols = [
        {'data': extra_data['types'], 's1_letter': 'B'},
        {'data': extra_data['categories'], 's1_letter': 'E'},
        {'data': extra_data['sale_categories'], 's1_letter': 'F'},
        {'data': extra_data['brands'], 's1_letter': 'G'},
        {'data': extra_data['colors'], 's1_letter': 'H'},
        {'data': extra_data['units'], 's1_letter': ['L', 'M']},
        {'data': extra_data['product_types'], 's1_letter': 'N'},
        {'data': extra_data['objectives'], 's1_letter': 'O'},
        {'data': extra_data['channels'], 's1_letter': 'X'},
        {'data': extra_data['editing_status'], 's1_letter': ''},
        {'data': extra_data['selling_status'], 's1_letter': ''},
        {'data': extra_data['product_units'], 's1_letter': ''},
        {'data': extra_data['tax_in'], 's1_letter': 'V'},
        {'data': extra_data['tax_out'], 's1_letter': 'W'}
    ]
    for key, value in enumerate(default_cols):
        s2_letter = get_column_letter(key + 1)
        for v, c in zip(
                value['data'],
                ws['{}3'.format(s2_letter):'{}{}'.format(s2_letter, len(value['data']) + 2)]):
            c[0].value = v

    # Write attribute options to DuLieuMau sheet
    if not export_common:
        attr_col_index = start_col
        for attribute in attributes_data.get('attributes', []):
            if attribute.value_type in ['selection', 'multiple_select']:
                s2_letter = get_column_letter(attr_col_index)
                # Header
                cell = ws['{}2'.format(s2_letter)]
                cell.value = attribute.name
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=bg_color, fill_type='solid')
                # Content
                options = [option.value for option in attribute.options]
                for v, c in zip(options, ws['{}3'.format(s2_letter):'{}{}'.format(
                        s2_letter, len(options) + 2)]):
                    c[0].value = v
                # Increase index
                attr_col_index += 1
